# from __future__ import annotations

# import json
# import re
# from dataclasses import dataclass
# from html import unescape as html_unescape
# from typing import Any, Dict, List, Optional, Tuple


# # --------------------------- Normalization ---------------------------

# _ws_re = re.compile(r"\s+")
# _nonprint_re = re.compile(r"[^\x09\x0A\x0D\x20-\x7E\u00A0-\uFFFF]")


# def norm(s: str) -> str:
#     s = html_unescape(s)
#     s = s.replace("\u00a0", " ")
#     s = _nonprint_re.sub(" ", s)
#     s = _ws_re.sub(" ", s.strip())
#     return s.upper()


# def is_leaf_value(v: Any) -> bool:
#     return isinstance(v, (str, int, float, bool))


# def leaf_to_string(v: Any) -> str:
#     if isinstance(v, bool):
#         return "TRUE" if v else "FALSE"
#     return str(v)


# def _to_int_page(x: Any) -> Optional[int]:
#     """Parse page_number from int or digit-string."""
#     if isinstance(x, int):
#         return x
#     if isinstance(x, str):
#         s = x.strip()
#         if s.isdigit():
#             return int(s)
#     return None


# # --------------------------- GT flattening (page-aware) ---------------------------

# @dataclass
# class GTLeaf:
#     path: str
#     value: str
#     page_number: Optional[int] = None


# def flatten_gt(obj: Any, prefix: str = "", page_ctx: Optional[int] = None) -> List[GTLeaf]:
#     out: List[GTLeaf] = []

#     if isinstance(obj, dict):
#         # Accept int or digit-string page_number
#         pn = _to_int_page(obj.get("page_number"))
#         if pn is not None:
#             page_ctx = pn

#         for k, v in obj.items():
#             new_prefix = f"{prefix}.{k}" if prefix else k
#             out.extend(flatten_gt(v, new_prefix, page_ctx))

#     elif isinstance(obj, list):
#         for i, item in enumerate(obj):
#             new_prefix = f"{prefix}[{i}]"
#             branch_ctx = page_ctx
#             if isinstance(item, dict):
#                 pn = _to_int_page(item.get("page_number"))
#                 if pn is not None:
#                     branch_ctx = pn
#             out.extend(flatten_gt(item, new_prefix, branch_ctx))

#     else:
#         if obj is None:
#             return out
#         if is_leaf_value(obj):
#             out.append(GTLeaf(path=prefix, value=leaf_to_string(obj), page_number=page_ctx))

#     return out


# # --------------------------- Safe load (JSON or raw text) ---------------------------

# def safe_load_json_bytes(b: bytes) -> Any:
#     txt = b.decode("utf-8", errors="replace")
#     try:
#         return json.loads(txt)
#     except Exception as e:
#         return {"raw_text": txt, "_parse_error": str(e)}


# # --------------------------- Candidate parsing ---------------------------

# def looks_like_merged_glmocr(d: Any) -> bool:
#     """
#     GLM merged (typical):
#       {"pages":[{"page_number":1,"data":{"markdown_result":...,"json_result":[[...]...]}}]}
#     Also accept variants where markdown_result/json_result are directly under page dict.
#     """
#     if not isinstance(d, dict):
#         return False
#     pages = d.get("pages")
#     if not isinstance(pages, list):
#         return False

#     for p in pages:
#         if not isinstance(p, dict):
#             continue
#         if isinstance(p.get("data"), dict):
#             return True
#         # variant: page may directly contain markdown_result/json_result
#         if isinstance(p.get("markdown_result"), str) or isinstance(p.get("json_result"), list):
#             return True
#     return False


# def collect_strings_from_ocr_data(data: Dict[str, Any]) -> List[str]:
#     parts: List[str] = []

#     md = data.get("markdown_result")
#     if isinstance(md, str) and md.strip():
#         parts.append(md)

#     jr = data.get("json_result")
#     if isinstance(jr, list):
#         for block in jr:
#             if isinstance(block, list):
#                 for region in block:
#                     if isinstance(region, dict):
#                         c = region.get("content")
#                         if isinstance(c, str) and c.strip():
#                             parts.append(c)

#     return parts


# def _gather_all_strings_generic(obj: Any, out: List[str]) -> None:
#     if obj is None:
#         return
#     if isinstance(obj, dict):
#         for k, v in obj.items():
#             if k in ("raw_text", "content", "text", "markdown", "markdown_result") and isinstance(v, str) and v.strip():
#                 out.append(v)
#             _gather_all_strings_generic(v, out)
#     elif isinstance(obj, list):
#         for x in obj:
#             _gather_all_strings_generic(x, out)
#     else:
#         if is_leaf_value(obj):
#             s = leaf_to_string(obj)
#             if s.strip():
#                 out.append(s)


# def build_page_texts_from_glm_merged(merged: Dict[str, Any]) -> Tuple[Dict[int, str], str]:
#     page_texts: Dict[int, str] = {}
#     all_parts: List[str] = []

#     for p in merged.get("pages", []):
#         if not isinstance(p, dict):
#             continue

#         pn = _to_int_page(p.get("page_number"))
#         if pn is None:
#             continue

#         # Typical: page["data"] is dict; Variant: page itself holds markdown_result/json_result
#         data = p.get("data") if isinstance(p.get("data"), dict) else p
#         if not isinstance(data, dict):
#             continue

#         parts = collect_strings_from_ocr_data(data)
#         raw = "\n".join(parts)
#         page_texts[pn] = norm(raw)
#         all_parts.append(raw)

#     full = norm("\n".join(all_parts))
#     return page_texts, full


# def build_page_texts_from_generic_pages(candidate: Dict[str, Any]) -> Optional[Tuple[Dict[int, str], str]]:
#     """
#     Generic support:
#       {"pages":[{"page_number":1, ... text fields ...}, ...]}
#     We recursively collect all strings inside each page object.
#     """
#     pages = candidate.get("pages")
#     if not isinstance(pages, list):
#         return None

#     page_texts: Dict[int, str] = {}
#     all_parts: List[str] = []
#     found_any = False

#     for p in pages:
#         if not isinstance(p, dict):
#             continue

#         pn = _to_int_page(p.get("page_number"))
#         if pn is None:
#             continue

#         tmp: List[str] = []
#         _gather_all_strings_generic(p, tmp)
#         raw = "\n".join(tmp).strip()
#         if raw:
#             found_any = True
#             page_texts[pn] = norm(raw)
#             all_parts.append(raw)

#     if not found_any:
#         return None
#     return page_texts, norm("\n".join(all_parts))


# def build_candidate_index(candidate: Any) -> Tuple[Dict[int, str], str]:
#     """
#     Returns (page_texts_norm, full_text_norm).
#     - If candidate has usable pages -> page_texts_norm populated
#     - Else -> page_texts_norm empty and full_text_norm from all strings / raw text
#     """
#     # Raw text wrapper (invalid JSON)
#     if isinstance(candidate, dict) and isinstance(candidate.get("raw_text"), str):
#         return {}, norm(candidate.get("raw_text", ""))

#     # GLM merged
#     if looks_like_merged_glmocr(candidate):
#         return build_page_texts_from_glm_merged(candidate)  # type: ignore[arg-type]

#     # Generic pages[]
#     if isinstance(candidate, dict):
#         gen = build_page_texts_from_generic_pages(candidate)
#         if gen is not None:
#             return gen

#     # Fully generic: dict/list/scalar -> gather everything
#     parts: List[str] = []
#     _gather_all_strings_generic(candidate, parts)
#     return {}, norm("\n".join(parts))


# # --------------------------- Evaluation ---------------------------

# @dataclass
# class MatchResult:
#     path: str
#     value: str
#     page_number: Optional[int]
#     matched: bool


# def _sort_scope_keys(keys: List[str]) -> List[str]:
#     """
#     Numeric pages in numeric order, DOC last.
#     """
#     nums = []
#     has_doc = False
#     others = []
#     for k in keys:
#         if k == "DOC":
#             has_doc = True
#         elif isinstance(k, str) and k.isdigit():
#             nums.append(int(k))
#         else:
#             others.append(k)

#     out = [str(x) for x in sorted(nums)]
#     out.extend(sorted(others))
#     if has_doc:
#         out.append("DOC")
#     return out


# def evaluate_overlap(gt: Any, candidate: Any) -> Dict[str, Any]:
#     gt_leaves = flatten_gt(gt)
#     page_texts_norm, full_text_norm = build_candidate_index(candidate)

#     results: List[MatchResult] = []
#     for leaf in gt_leaves:
#         v_norm = norm(leaf.value)
#         if not v_norm:
#             continue

#         # If GT has page_number and candidate has that page extracted, use it
#         if leaf.page_number is not None and leaf.page_number in page_texts_norm:
#             hay = page_texts_norm[leaf.page_number]
#         else:
#             hay = full_text_norm

#         matched = v_norm in hay
#         results.append(MatchResult(leaf.path, leaf.value, leaf.page_number, matched))

#     by_scope: Dict[str, Dict[str, int]] = {}
#     for r in results:
#         key = str(r.page_number) if r.page_number is not None else "DOC"
#         if key not in by_scope:
#             by_scope[key] = {"total": 0, "hit": 0}
#         by_scope[key]["total"] += 1
#         by_scope[key]["hit"] += 1 if r.matched else 0

#     total = len(results)
#     hit = sum(1 for r in results if r.matched)
#     recall = (hit / total) if total else 0.0

#     # stable ordering
#     ordered_keys = _sort_scope_keys(list(by_scope.keys()))

#     report = {
#         "metric": "substring_overlap_recall",
#         "total_fields_checked": total,
#         "matched_fields": hit,
#         "recall": round(recall, 6),
#         "breakdown_by_page": {
#             k: {
#                 "total": by_scope[k]["total"],
#                 "hit": by_scope[k]["hit"],
#                 "recall": round((by_scope[k]["hit"] / by_scope[k]["total"]) if by_scope[k]["total"] else 0.0, 6),
#             }
#             for k in ordered_keys
#         },
#     }
#     return report

#############################################################################################################################################################################

# from __future__ import annotations

# import io
# import json
# import math
# import re
# from dataclasses import dataclass
# from html import unescape as html_unescape
# from typing import Any, Dict, List, Optional, Tuple

# import pandas as pd
# from openpyxl import Workbook
# from openpyxl.styles import Alignment, Font, PatternFill


# # --------------------------- Normalization ---------------------------

# _ws_re = re.compile(r"\s+")
# _nonprint_re = re.compile(r"[^\x09\x0A\x0D\x20-\x7E\u00A0-\uFFFF]")


# def norm(s: str) -> str:
#     s = html_unescape(s)
#     s = s.replace("\u00a0", " ")
#     s = _nonprint_re.sub(" ", s)
#     s = _ws_re.sub(" ", s.strip())
#     return s.upper()


# def is_leaf_value(v: Any) -> bool:
#     return isinstance(v, (str, int, float, bool))


# def leaf_to_string(v: Any) -> str:
#     if isinstance(v, bool):
#         return "TRUE" if v else "FALSE"
#     return str(v)


# def _to_int_page(x: Any) -> Optional[int]:
#     """Parse page_number from int or digit-string."""
#     if isinstance(x, int):
#         return x
#     if isinstance(x, str):
#         s = x.strip()
#         if s.isdigit():
#             return int(s)
#     return None


# def _is_empty_cell(x: Any) -> bool:
#     if x is None:
#         return True
#     if isinstance(x, float) and (math.isnan(x) or pd.isna(x)):
#         return True
#     s = str(x).strip()
#     return s == "" or s.lower() in {"nan", "none", "null"}


# def _key_piece(s: Any) -> str:
#     """Normalize a hierarchy segment into snake-ish lowercase."""
#     t = str(s).strip()
#     t = re.sub(r"\s+", " ", t)
#     t = t.replace("&", "and")
#     t = re.sub(r"[^A-Za-z0-9]+", "_", t)
#     t = re.sub(r"_+", "_", t).strip("_")
#     return t.lower()


# # --------------------------- GT flattening (page-aware) ---------------------------

# @dataclass
# class GTLeaf:
#     path: str
#     value: str
#     page_number: Optional[int] = None


# def flatten_gt(obj: Any, prefix: str = "", page_ctx: Optional[int] = None) -> List[GTLeaf]:
#     out: List[GTLeaf] = []

#     if isinstance(obj, dict):
#         pn = _to_int_page(obj.get("page_number"))
#         if pn is not None:
#             page_ctx = pn

#         for k, v in obj.items():
#             new_prefix = f"{prefix}.{k}" if prefix else k
#             out.extend(flatten_gt(v, new_prefix, page_ctx))

#     elif isinstance(obj, list):
#         for i, item in enumerate(obj):
#             new_prefix = f"{prefix}[{i}]"
#             branch_ctx = page_ctx
#             if isinstance(item, dict):
#                 pn = _to_int_page(item.get("page_number"))
#                 if pn is not None:
#                     branch_ctx = pn
#             out.extend(flatten_gt(item, new_prefix, branch_ctx))

#     else:
#         if obj is None:
#             return out
#         if is_leaf_value(obj):
#             out.append(GTLeaf(path=prefix, value=leaf_to_string(obj), page_number=page_ctx))

#     return out


# # --------------------------- Safe load (JSON / TXT / Excel) ---------------------------

# def safe_load_json_bytes(b: bytes) -> Any:
#     txt = b.decode("utf-8", errors="replace")
#     try:
#         return json.loads(txt)
#     except Exception as e:
#         return {"raw_text": txt, "_parse_error": str(e)}


# def safe_load_text_bytes(b: bytes) -> Any:
#     txt = b.decode("utf-8", errors="replace")
#     return {"raw_text": txt}


# def safe_load_excel_bytes(b: bytes) -> pd.DataFrame:
#     bio = io.BytesIO(b)
#     return pd.read_excel(bio)


# # --------------------------- Candidate parsing (for overlap recall) ---------------------------

# def looks_like_merged_glmocr(d: Any) -> bool:
#     """
#     GLM merged (typical):
#       {"pages":[{"page_number":1,"data":{"markdown_result":...,"json_result":[[...]...]}}]}
#     Also accept variants where markdown_result/json_result are directly under page dict.
#     """
#     if not isinstance(d, dict):
#         return False
#     pages = d.get("pages")
#     if not isinstance(pages, list):
#         return False

#     for p in pages:
#         if not isinstance(p, dict):
#             continue
#         if isinstance(p.get("data"), dict):
#             return True
#         if isinstance(p.get("markdown_result"), str) or isinstance(p.get("json_result"), list):
#             return True
#     return False


# def collect_strings_from_ocr_data(data: Dict[str, Any]) -> List[str]:
#     parts: List[str] = []

#     md = data.get("markdown_result")
#     if isinstance(md, str) and md.strip():
#         parts.append(md)

#     jr = data.get("json_result")
#     if isinstance(jr, list):
#         for block in jr:
#             if isinstance(block, list):
#                 for region in block:
#                     if isinstance(region, dict):
#                         c = region.get("content")
#                         if isinstance(c, str) and c.strip():
#                             parts.append(c)

#     return parts


# def _gather_all_strings_generic(obj: Any, out: List[str]) -> None:
#     if obj is None:
#         return
#     if isinstance(obj, dict):
#         for k, v in obj.items():
#             if k in ("raw_text", "content", "text", "markdown", "markdown_result") and isinstance(v, str) and v.strip():
#                 out.append(v)
#             _gather_all_strings_generic(v, out)
#     elif isinstance(obj, list):
#         for x in obj:
#             _gather_all_strings_generic(x, out)
#     else:
#         if is_leaf_value(obj):
#             s = leaf_to_string(obj)
#             if s.strip():
#                 out.append(s)


# def build_page_texts_from_glm_merged(merged: Dict[str, Any]) -> Tuple[Dict[int, str], str]:
#     page_texts: Dict[int, str] = {}
#     all_parts: List[str] = []

#     for p in merged.get("pages", []):
#         if not isinstance(p, dict):
#             continue

#         pn = _to_int_page(p.get("page_number"))
#         if pn is None:
#             continue

#         data = p.get("data") if isinstance(p.get("data"), dict) else p
#         if not isinstance(data, dict):
#             continue

#         parts = collect_strings_from_ocr_data(data)
#         raw = "\n".join(parts)
#         page_texts[pn] = norm(raw)
#         all_parts.append(raw)

#     full = norm("\n".join(all_parts))
#     return page_texts, full


# def build_page_texts_from_generic_pages(candidate: Dict[str, Any]) -> Optional[Tuple[Dict[int, str], str]]:
#     """
#     Generic support:
#       {"pages":[{"page_number":1, ... text fields ...}, ...]}
#     """
#     pages = candidate.get("pages")
#     if not isinstance(pages, list):
#         return None

#     page_texts: Dict[int, str] = {}
#     all_parts: List[str] = []
#     found_any = False

#     for p in pages:
#         if not isinstance(p, dict):
#             continue

#         pn = _to_int_page(p.get("page_number"))
#         if pn is None:
#             continue

#         tmp: List[str] = []
#         _gather_all_strings_generic(p, tmp)
#         raw = "\n".join(tmp).strip()
#         if raw:
#             found_any = True
#             page_texts[pn] = norm(raw)
#             all_parts.append(raw)

#     if not found_any:
#         return None
#     return page_texts, norm("\n".join(all_parts))


# def build_candidate_index(candidate: Any) -> Tuple[Dict[int, str], str]:
#     """
#     Returns (page_texts_norm, full_text_norm).
#     - If candidate has usable pages -> page_texts_norm populated
#     - Else -> page_texts_norm empty and full_text_norm from all strings / raw text
#     """
#     if isinstance(candidate, dict) and isinstance(candidate.get("raw_text"), str):
#         return {}, norm(candidate.get("raw_text", ""))

#     if looks_like_merged_glmocr(candidate):
#         return build_page_texts_from_glm_merged(candidate)  # type: ignore[arg-type]

#     if isinstance(candidate, dict):
#         gen = build_page_texts_from_generic_pages(candidate)
#         if gen is not None:
#             return gen

#     parts: List[str] = []
#     _gather_all_strings_generic(candidate, parts)
#     return {}, norm("\n".join(parts))


# # --------------------------- Evaluation (Overlap Recall) ---------------------------

# @dataclass
# class MatchResult:
#     path: str
#     value: str
#     page_number: Optional[int]
#     matched: bool


# def _sort_scope_keys(keys: List[str]) -> List[str]:
#     """Numeric pages in numeric order, DOC last."""
#     nums = []
#     has_doc = False
#     others = []
#     for k in keys:
#         if k == "DOC":
#             has_doc = True
#         elif isinstance(k, str) and k.isdigit():
#             nums.append(int(k))
#         else:
#             others.append(k)

#     out = [str(x) for x in sorted(nums)]
#     out.extend(sorted(others))
#     if has_doc:
#         out.append("DOC")
#     return out


# def evaluate_overlap_from_leaves(gt_leaves: List[GTLeaf], candidate: Any) -> Dict[str, Any]:
#     page_texts_norm, full_text_norm = build_candidate_index(candidate)

#     results: List[MatchResult] = []
#     for leaf in gt_leaves:
#         v_norm = norm(leaf.value)
#         if not v_norm:
#             continue

#         if leaf.page_number is not None and leaf.page_number in page_texts_norm:
#             hay = page_texts_norm[leaf.page_number]
#         else:
#             hay = full_text_norm

#         matched = v_norm in hay
#         results.append(MatchResult(leaf.path, leaf.value, leaf.page_number, matched))

#     by_scope: Dict[str, Dict[str, int]] = {}
#     for r in results:
#         key = str(r.page_number) if r.page_number is not None else "DOC"
#         if key not in by_scope:
#             by_scope[key] = {"total": 0, "hit": 0}
#         by_scope[key]["total"] += 1
#         by_scope[key]["hit"] += 1 if r.matched else 0

#     total = len(results)
#     hit = sum(1 for r in results if r.matched)
#     recall = (hit / total) if total else 0.0

#     ordered_keys = _sort_scope_keys(list(by_scope.keys()))

#     report = {
#         "metric": "substring_overlap_recall",
#         "total_fields_checked": total,
#         "matched_fields": hit,
#         "recall": round(recall, 6),
#         "breakdown_by_page": {
#             k: {
#                 "total": by_scope[k]["total"],
#                 "hit": by_scope[k]["hit"],
#                 "recall": round((by_scope[k]["hit"] / by_scope[k]["total"]) if by_scope[k]["total"] else 0.0, 6),
#             }
#             for k in ordered_keys
#         },
#     }
#     return report


# def evaluate_overlap(gt: Any, candidate: Any) -> Dict[str, Any]:
#     """Backwards-compatible: GT JSON -> flatten -> evaluate."""
#     gt_leaves = flatten_gt(gt)
#     return evaluate_overlap_from_leaves(gt_leaves, candidate)


# # --------------------------- Excel -> Leaves (Doc-type aware) ---------------------------

# def _canon_section(doc_type: str, section: Any) -> str:
#     s = str(section).strip().lower()
#     if doc_type == "shipping_bill":
#         if "shipping bill summary" in s:
#             return "shippingbillsummary"
#         if s == "header":
#             return "header"
#     if doc_type == "purchase_order":
#         if "purchase" in s and "order" in s:
#             return "purchaseorder"
#     return _key_piece(section)


# def shipping_bill_gt_excel_to_leaves(df: pd.DataFrame) -> List[GTLeaf]:
#     """
#     Matches the MI-2351-6827 shipping bill GT excel style.
#     Uses:
#       - Case A: section + group + key(col2) -> value(col3)
#       - Case B: section + key(col1) -> value(col2)
#     """
#     cols = list(df.columns)
#     if len(cols) < 4:
#         return []

#     c0, c1, c2, c3 = cols[0], cols[1], cols[2], cols[3]

#     ff = df.copy()
#     ff[c0] = ff[c0].ffill()
#     ff[c1] = ff[c1].ffill()

#     out: List[GTLeaf] = []
#     for i, row in df.iterrows():
#         section = ff.at[i, c0]
#         group = ff.at[i, c1]
#         r1, r2, r3 = row[c1], row[c2], row[c3]

#         # Case A
#         if (not _is_empty_cell(r2)) and (not _is_empty_cell(r3)):
#             path = ".".join(
#                 [p for p in [_canon_section("shipping_bill", section), _key_piece(group), _key_piece(r2)] if p]
#             )
#             out.append(GTLeaf(path=path, value=leaf_to_string(r3), page_number=None))
#             continue

#         # Case B
#         if (not _is_empty_cell(r1)) and (not _is_empty_cell(r2)):
#             path = ".".join([p for p in [_canon_section("shipping_bill", section), _key_piece(r1)] if p])
#             out.append(GTLeaf(path=path, value=leaf_to_string(r2), page_number=None))
#             continue

#     return out


# def purchase_order_gt_excel_to_leaves(df: pd.DataFrame) -> List[GTLeaf]:
#     """
#     Handles common PO GT excel layouts:
#       A) columns include "Key"/"Field" and "Value"
#       B) L1..Ln + Value
#       C) fallback first 2 columns key/value
#     """
#     cols = list(df.columns)
#     if not cols:
#         return []

#     # A) explicit Key/Value
#     low = [str(c).lower() for c in cols]
#     key_col = None
#     val_col = None
#     for i, c in enumerate(low):
#         if c in {"key", "field", "path"}:
#             key_col = cols[i]
#         if c in {"value", "val"}:
#             val_col = cols[i]
#     if key_col is not None and val_col is not None:
#         out: List[GTLeaf] = []
#         for _, r in df.iterrows():
#             k = r[key_col]
#             v = r[val_col]
#             if _is_empty_cell(k) or _is_empty_cell(v):
#                 continue
#             out.append(GTLeaf(path=_key_piece(k).replace("_", "."), value=leaf_to_string(v), page_number=None))
#         return out

#     # B) L1..Ln + Value
#     lcols = [c for c in cols if str(c).lower().startswith("l")]
#     if "Value" in cols and lcols:
#         out = []
#         for _, r in df.iterrows():
#             parts = []
#             for c in lcols:
#                 if not _is_empty_cell(r[c]):
#                     parts.append(_key_piece(r[c]))
#             if not parts:
#                 continue
#             v = r["Value"]
#             if _is_empty_cell(v):
#                 continue
#             out.append(GTLeaf(path=".".join(parts), value=leaf_to_string(v), page_number=None))
#         return out

#     # C) fallback: first 2 columns
#     if len(cols) >= 2:
#         out = []
#         for _, r in df.iterrows():
#             k = r[cols[0]]
#             v = r[cols[1]]
#             if _is_empty_cell(k) or _is_empty_cell(v):
#                 continue
#             out.append(GTLeaf(path=_key_piece(k).replace("_", "."), value=leaf_to_string(v), page_number=None))
#         return out

#     return []


# def pred_excel_to_leaves(df: pd.DataFrame) -> List[GTLeaf]:
#     """
#     Pred excel formats supported:
#       - Path + Value
#       - L1..Ln + Value
#     """
#     cols = list(df.columns)
#     if not cols:
#         return []

#     if "Path" in cols and "Value" in cols:
#         out: List[GTLeaf] = []
#         for _, r in df.iterrows():
#             p = r["Path"]
#             v = r["Value"]
#             if _is_empty_cell(p):
#                 continue
#             # keep a consistent dot-hierarchy (convert underscores to dots only lightly)
#             p_norm = _key_piece(p).replace("_", ".")
#             out.append(GTLeaf(path=p_norm, value=leaf_to_string(v) if not _is_empty_cell(v) else "", page_number=None))
#         return out

#     lcols = [c for c in cols if str(c).lower().startswith("l")]
#     if "Value" in cols and lcols:
#         out = []
#         for _, r in df.iterrows():
#             parts = []
#             for c in lcols:
#                 if not _is_empty_cell(r[c]):
#                     parts.append(_key_piece(r[c]))
#             if not parts:
#                 continue
#             v = r["Value"]
#             out.append(GTLeaf(path=".".join(parts), value=leaf_to_string(v) if not _is_empty_cell(v) else "", page_number=None))
#         return out

#     return []


# def parse_gt(doc_type: str, fmt: str, payload: Any) -> List[GTLeaf]:
#     if fmt == "json":
#         return flatten_gt(payload)
#     # excel
#     assert isinstance(payload, pd.DataFrame)
#     if doc_type == "shipping_bill":
#         return shipping_bill_gt_excel_to_leaves(payload)
#     if doc_type == "purchase_order":
#         return purchase_order_gt_excel_to_leaves(payload)
#     return []


# def parse_pred(fmt: str, payload: Any) -> List[GTLeaf]:
#     if fmt == "json":
#         return flatten_gt(payload)
#     assert isinstance(payload, pd.DataFrame)
#     return pred_excel_to_leaves(payload)


# # --------------------------- Alignment Excel (Color-coded) ---------------------------

# def _sig(s: str) -> str:
#     toks = re.split(r"[.\[\]_/:\-()\s]+", s or "")
#     normt = []
#     for t in toks:
#         t = t.strip()
#         if not t:
#             continue
#         normt.append(_key_piece(t))
#     normt = [x for x in normt if x]
#     return "|".join(sorted(set(normt)))


# def build_alignment_xlsx_bytes(gt_leaves: List[GTLeaf], pred_leaves: List[GTLeaf]) -> bytes:
#     """
#     Output columns:
#       GT_Key(Hierarchy), Pred_Key(Hierarchy), GT_Value, Pred_Value, Status
#     Row colors:
#       MATCH=green, MISMATCH=yellow, MISSING_KEY or MISSING_VALUE=red
#     """

#     # index pred by signatures
#     pred_idx: Dict[str, List[int]] = {}
#     for i, pl in enumerate(pred_leaves):
#         p = pl.path or ""
#         segs = [x for x in p.split(".") if x]
#         sigs = set()
#         sigs.add(_sig(p))
#         if segs:
#             sigs.add(_sig(segs[-1]))
#         if len(segs) >= 2:
#             sigs.add(_sig(".".join(segs[-2:])))
#         for s in sigs:
#             if not s:
#                 continue
#             pred_idx.setdefault(s, []).append(i)

#     used: set[int] = set()

#     def _nv(x: str) -> str:
#         return norm(str(x)) if x is not None else ""

#     rows: List[Tuple[str, str, str, str, str]] = []
#     for gl in gt_leaves:
#         gk = gl.path or ""
#         gv = leaf_to_string(gl.value)

#         g_last = gk.split(".")[-1] if gk else ""
#         gsig = _sig(g_last) or _sig(gk)

#         cands = pred_idx.get(gsig, [])
#         best_i: Optional[int] = None

#         if cands:
#             best_score = -1e9
#             for ci in cands:
#                 score = 0.0
#                 if ci in used:
#                     score -= 0.1
#                 pv = pred_leaves[ci].value
#                 if _nv(pv) == _nv(gv) and _nv(pv) != "":
#                     score += 2.0
#                 score -= abs(len(pred_leaves[ci].path or "") - len(gk)) / 2000.0
#                 if score > best_score:
#                     best_score = score
#                     best_i = ci

#         if best_i is None:
#             rows.append((gk, "", gv, "", "MISSING_KEY"))
#         else:
#             used.add(best_i)
#             pk = pred_leaves[best_i].path or ""
#             pv = pred_leaves[best_i].value or ""
#             if _is_empty_cell(pv):
#                 st = "MISSING_VALUE"
#             else:
#                 st = "MATCH" if _nv(pv) == _nv(gv) else "MISMATCH"
#             rows.append((gk, pk, gv, leaf_to_string(pv), st))

#     wb = Workbook()
#     ws = wb.active
#     ws.title = "GT_vs_Pred"

#     ws.append(["GT_Key(Hierarchy)", "Pred_Key(Hierarchy)", "GT_Value", "Pred_Value", "Status"])

#     header_fill = PatternFill("solid", fgColor="1F2937")
#     header_font = Font(color="FFFFFF", bold=True)
#     center = Alignment(vertical="center", wrap_text=True)

#     for c in ws[1]:
#         c.fill = header_fill
#         c.font = header_font
#         c.alignment = center

#     green = PatternFill("solid", fgColor="C6EFCE")
#     yellow = PatternFill("solid", fgColor="FFEB9C")
#     red = PatternFill("solid", fgColor="FFC7CE")

#     for r in rows:
#         ws.append(list(r))

#     for i in range(2, ws.max_row + 1):
#         st = ws.cell(row=i, column=5).value
#         fill = green if st == "MATCH" else yellow if st == "MISMATCH" else red
#         for j in range(1, 6):
#             ws.cell(row=i, column=j).fill = fill
#             ws.cell(row=i, column=j).alignment = center

#     ws.freeze_panes = "A2"
#     ws.column_dimensions["A"].width = 55
#     ws.column_dimensions["B"].width = 55
#     ws.column_dimensions["C"].width = 25
#     ws.column_dimensions["D"].width = 25
#     ws.column_dimensions["E"].width = 14

#     # Legend
#     ws2 = wb.create_sheet("Legend")
#     ws2.append(["Color", "Status", "Meaning"])
#     for c in ws2[1]:
#         c.fill = header_fill
#         c.font = header_font
#         c.alignment = center

#     legend_rows = [
#         ("Green", "MATCH", "GT value equals Pred value (normalized)"),
#         ("Yellow", "MISMATCH", "Key matched but values differ"),
#         ("Red", "MISSING_KEY / MISSING_VALUE", "No matched pred key OR pred value empty"),
#     ]
#     for row in legend_rows:
#         ws2.append(list(row))

#     for i in range(2, 5):
#         status = ws2.cell(row=i, column=2).value
#         fill = green if status == "MATCH" else yellow if status == "MISMATCH" else red
#         for j in range(1, 4):
#             ws2.cell(row=i, column=j).fill = fill
#             ws2.cell(row=i, column=j).alignment = center

#     ws2.column_dimensions["A"].width = 12
#     ws2.column_dimensions["B"].width = 28
#     ws2.column_dimensions["C"].width = 54
#     ws2.freeze_panes = "A2"

#     bio = io.BytesIO()
#     wb.save(bio)
#     return bio.getvalue()



# /home/mtq3kor/aman/GLM/glm-ocr/eval_backend.py
# /home/mtq3kor/aman/GLM/glm-ocr/eval_backend.py
from __future__ import annotations

import io
import json
import math
import re
from dataclasses import dataclass
from html import unescape as html_unescape
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


# ============================================================
# Normalization (DISPLAY + MATCHING)
# ============================================================

_ws_re = re.compile(r"\s+")
_nonprint_re = re.compile(r"[^\x09\x0A\x0D\x20-\x7E\u00A0-\uFFFF]")

# keep only a-z0-9 . , - and spaces
_meaningless_special_re = re.compile(r"[^a-z0-9\.\,\-\s]+")

_dash_re = re.compile(r"[–—−]")  # unicode dashes -> "-"
_hyphen_space_re = re.compile(r"\s*-\s*")
_commas_re = re.compile(r"\s*,\s*")

_item_num_re = re.compile(r"\b(\d+)\b")


def norm(s: Any) -> str:
    """
    STRICT STRING normalization:
      - html unescape
      - normalize unicode dashes to "-"
      - lowercase
      - remove meaningless special chars (keep . , -)
      - collapse whitespace
      - REMOVE ALL SPACES finally
    """
    s = html_unescape(str(s))
    s = s.replace("\u00a0", " ")
    s = _nonprint_re.sub(" ", s)
    s = _dash_re.sub("-", s)

    s = s.lower()
    s = _meaningless_special_re.sub(" ", s)

    s = _commas_re.sub(", ", s)
    s = _hyphen_space_re.sub("-", s)

    s = _ws_re.sub(" ", s).strip()
    s = s.replace(" ", "")  # remove ALL spaces
    return s


def norm_numeric(x: Any) -> str:
    """
    STRICT NUMERIC normalization for DISPLAY + MATCH:
      - strips commas
      - float parse
      - integer-valued => int string (0.0 -> 0)
      - removes trailing zeros
      - fallback => norm()
    """
    raw = html_unescape(str(x)).strip()
    if raw == "":
        return ""
    t = raw.replace(",", "").strip()
    try:
        f = float(t)
        if abs(f - int(f)) < 1e-9:
            return str(int(f))
        out = f"{f}".rstrip("0").rstrip(".") if "." in f"{f}" else f"{f}"
        return out
    except Exception:
        return norm(raw)


def norm_loose(x: Any) -> str:
    """
    Looser normalization for comparison only.
    Removes commas and hyphens after standard normalization.
    """
    s = norm(x)
    s = s.replace(",", "")
    s = s.replace("-", "")
    return s


def _soft_text(x: Any) -> str:
    """
    Softer normalization for token-based similarity.
    Keeps spaces for tokenization.
    """
    x = html_unescape(str(x)).lower()
    x = x.replace("\u00a0", " ")
    x = _nonprint_re.sub(" ", x)
    x = _dash_re.sub("-", x)
    x = _meaningless_special_re.sub(" ", x)
    x = x.replace("\n", " ")
    x = _commas_re.sub(", ", x)
    x = _hyphen_space_re.sub("-", x)
    x = _ws_re.sub(" ", x).strip()
    return x


def _soft_tokens(x: Any) -> List[str]:
    s = _soft_text(x)
    s = s.replace(",", " ")
    return [t for t in s.split(" ") if t]


def is_numeric_like(x: Any) -> bool:
    """
    Generic numeric detector based on value itself, not field/path.
    """
    s = html_unescape(str(x)).strip()
    if s == "":
        return False

    s = s.replace(",", "")
    s = s.replace("\u00a0", "")
    s = s.strip()

    return bool(re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s))


def alpha_ratio(x: Any) -> float:
    """
    Fraction of alphabetic chars in normalized string.
    Helps separate text from codes/ids.
    """
    t = norm(x)
    if not t:
        return 0.0
    alpha = sum(ch.isalpha() for ch in t)
    return alpha / max(len(t), 1)


def looks_like_identifier(x: Any) -> bool:
    """
    Generic identifier/code detector.
    No path hardcoding.
    Examples caught:
      37440394640
      po12345
      ab-9981-x
      22a19c44
    """
    raw = html_unescape(str(x)).strip()
    t = norm(raw)
    if not t:
        return False

    # Pure numeric-like values behave like exact identifiers/numbers
    if is_numeric_like(raw):
        return True

    # Keep only compact single-token-ish strings as potential identifiers
    raw_has_space = bool(re.search(r"\s", raw.strip()))
    compact = not raw_has_space

    digit_ratio = sum(ch.isdigit() for ch in t) / max(len(t), 1)
    alnum_punct_only = bool(re.fullmatch(r"[a-z0-9\.\,\-]+", t))

    if compact and alnum_punct_only:
        if len(t) >= 6 and any(ch.isdigit() for ch in t):
            return True
        if digit_ratio >= 0.5 and len(t) >= 4:
            return True

    return False


def token_count_soft(x: Any) -> int:
    return len(_soft_tokens(x))


def is_text_like_for_fuzzy(x: Any) -> bool:
    """
    Value-level gate for allowing NEAR_MATCH.
    """
    t = norm(x)
    if not t:
        return False

    if len(t) < 4:
        return False

    if looks_like_identifier(x):
        return False

    if alpha_ratio(x) < 0.4:
        return False

    if token_count_soft(x) < 2:
        return False

    return True


def display_value(path: str, v: Any) -> str:
    """
    What gets WRITTEN into Excel cells (GT/My/GPT values).
    Uses value-based numeric detection to keep it generic across PDFs.
    """
    if v is None:
        return ""
    if is_numeric_like(v):
        return norm_numeric(v)
    return norm(v)


# ============================================================
# Similarity for NEAR_MATCH
# ============================================================

def jaccard_tokens(a: Any, b: Any) -> float:
    """
    Token Jaccard for NEAR_MATCH.
    Restricted to multi-token text-like values only.
    """
    A = _soft_tokens(a)
    B = _soft_tokens(b)

    if len(A) < 2 or len(B) < 2:
        return 0.0

    SA, SB = set(A), set(B)
    if not SA and not SB:
        return 1.0
    if not SA or not SB:
        return 0.0
    return len(SA & SB) / len(SA | SB)


def contains_norm(a: Any, b: Any) -> bool:
    """
    Containment check after norm() (spaces removed).
    Restricted to real text-like values only.
    """
    a2, b2 = norm(a), norm(b)
    if not a2 or not b2:
        return False

    if len(a2) < 4 or len(b2) < 4:
        return False

    if not is_text_like_for_fuzzy(a) or not is_text_like_for_fuzzy(b):
        return False

    return a2 in b2 or b2 in a2


# ============================================================
# Basic helpers
# ============================================================

def is_leaf_value(v: Any) -> bool:
    return isinstance(v, (str, int, float, bool))


def leaf_to_string(v: Any) -> str:
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return str(v)


def _to_int_page(x: Any) -> Optional[int]:
    if isinstance(x, int):
        return x
    if isinstance(x, str):
        s = x.strip()
        if s.isdigit():
            return int(s)
    return None


def _is_empty_cell(x: Any) -> bool:
    if x is None:
        return True
    if isinstance(x, float) and (math.isnan(x) or pd.isna(x)):
        return True
    s = str(x).strip()
    return s == "" or s.lower() in {"nan", "none", "null"}


def _key_piece(s: Any) -> str:
    t = str(s).strip()
    t = re.sub(r"\s+", " ", t)
    t = t.replace("&", "and")
    t = re.sub(r"[^A-Za-z0-9]+", "_", t)
    t = re.sub(r"_+", "_", t).strip("_")
    return t.lower()


# ============================================================
# Doc-type heuristics for JSON
# ============================================================

def _detect_doc_type_from_json(obj: Any, fallback: str = "shipping_bill") -> str:
    if not isinstance(obj, dict):
        return fallback
    sv = str(obj.get("schema_version") or "").strip().lower()
    if "purchase" in sv and "order" in sv:
        return "purchase_order"
    src = obj.get("source")
    if isinstance(src, dict):
        dt = str(src.get("document_type") or "").strip().lower()
        if dt in {"purchase_order", "shipping_bill"}:
            return dt
    return fallback


def _extract_item_id(x: Any) -> Optional[str]:
    if x is None:
        return None
    s = str(x).strip()
    if not s:
        return None
    m = _item_num_re.search(s)
    return m.group(1) if m else None


def _canon_po_details_key(k: str) -> Optional[str]:
    item_id = _extract_item_id(k)
    if item_id is None:
        return None
    return f"Details[ITEM={item_id}]"


# ============================================================
# GT flattening (page-aware + PO canonicalization)
# ============================================================

@dataclass
class GTLeaf:
    path: str
    value: str
    page_number: Optional[int] = None


def flatten_gt(
    obj: Any,
    prefix: str = "",
    page_ctx: Optional[int] = None,
    doc_type: Optional[str] = None,
) -> List[GTLeaf]:
    out: List[GTLeaf] = []

    if doc_type is None:
        doc_type = _detect_doc_type_from_json(obj, fallback="shipping_bill")

    if isinstance(obj, dict):
        pn = _to_int_page(obj.get("page_number"))
        if pn is not None:
            page_ctx = pn

        # PO: canonicalize Details.* keys like Item1 -> Details[ITEM=1]
        if doc_type == "purchase_order" and (prefix == "Details" or prefix.endswith(".Details")):
            for k, v in obj.items():
                ck = _canon_po_details_key(str(k))
                if ck is None:
                    new_prefix = f"{prefix}.{k}" if prefix else str(k)
                else:
                    new_prefix = ck
                out.extend(flatten_gt(v, new_prefix, page_ctx, doc_type))
            return out

        for k, v in obj.items():
            new_prefix = f"{prefix}.{k}" if prefix else str(k)
            out.extend(flatten_gt(v, new_prefix, page_ctx, doc_type))

    elif isinstance(obj, list):
        # PO: canonicalize RowWiseTable list items using ITEM id
        if doc_type == "purchase_order" and (prefix == "RowWiseTable" or prefix.endswith(".RowWiseTable")):
            for i, item in enumerate(obj):
                branch_ctx = page_ctx
                if isinstance(item, dict):
                    pn = _to_int_page(item.get("page_number"))
                    if pn is not None:
                        branch_ctx = pn

                    item_id = _extract_item_id(item.get("Item"))
                    if item_id is None:
                        item_id = _extract_item_id(item.get("ShipNo"))
                    if item_id is not None:
                        new_prefix = f"RowWiseTable[ITEM={item_id}]"
                    else:
                        new_prefix = f"{prefix}[{i}]"
                else:
                    new_prefix = f"{prefix}[{i}]"

                out.extend(flatten_gt(item, new_prefix, branch_ctx, doc_type))
            return out

        for i, item in enumerate(obj):
            new_prefix = f"{prefix}[{i}]"
            branch_ctx = page_ctx
            if isinstance(item, dict):
                pn = _to_int_page(item.get("page_number"))
                if pn is not None:
                    branch_ctx = pn
            out.extend(flatten_gt(item, new_prefix, branch_ctx, doc_type))

    else:
        if obj is None:
            return out
        if is_leaf_value(obj):
            out.append(GTLeaf(path=prefix, value=leaf_to_string(obj), page_number=page_ctx))

    return out


# ============================================================
# Safe load
# ============================================================

def safe_load_json_bytes(b: bytes) -> Any:
    txt = b.decode("utf-8", errors="replace")
    try:
        return json.loads(txt)
    except Exception as e:
        return {"raw_text": txt, "_parse_error": str(e)}


def safe_load_text_bytes(b: bytes) -> Any:
    txt = b.decode("utf-8", errors="replace")
    return {"raw_text": txt}


def safe_load_excel_bytes(b: bytes) -> pd.DataFrame:
    bio = io.BytesIO(b)
    return pd.read_excel(bio)


# ============================================================
# Candidate parsing (overlap recall)
# ============================================================

def looks_like_merged_glmocr(d: Any) -> bool:
    if not isinstance(d, dict):
        return False
    pages = d.get("pages")
    if not isinstance(pages, list):
        return False

    for p in pages:
        if not isinstance(p, dict):
            continue
        if isinstance(p.get("data"), dict):
            return True
        if isinstance(p.get("markdown_result"), str) or isinstance(p.get("json_result"), list):
            return True
    return False


def collect_strings_from_ocr_data(data: Dict[str, Any]) -> List[str]:
    parts: List[str] = []

    md = data.get("markdown_result")
    if isinstance(md, str) and md.strip():
        parts.append(md)

    jr = data.get("json_result")
    if isinstance(jr, list):
        for block in jr:
            if isinstance(block, list):
                for region in block:
                    if isinstance(region, dict):
                        c = region.get("content")
                        if isinstance(c, str) and c.strip():
                            parts.append(c)

    return parts


def _gather_all_strings_generic(obj: Any, out: List[str]) -> None:
    if obj is None:
        return
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k in ("raw_text", "content", "text", "markdown", "markdown_result") and isinstance(v, str) and v.strip():
                out.append(v)
            _gather_all_strings_generic(v, out)
    elif isinstance(obj, list):
        for x in obj:
            _gather_all_strings_generic(x, out)
    else:
        if is_leaf_value(obj):
            s = leaf_to_string(obj)
            if s.strip():
                out.append(s)


def build_page_texts_from_glm_merged(merged: Dict[str, Any]) -> Tuple[Dict[int, str], str]:
    page_texts: Dict[int, str] = {}
    all_parts: List[str] = []

    for p in merged.get("pages", []):
        if not isinstance(p, dict):
            continue

        pn = _to_int_page(p.get("page_number"))
        if pn is None:
            continue

        data = p.get("data") if isinstance(p.get("data"), dict) else p
        if not isinstance(data, dict):
            continue

        parts = collect_strings_from_ocr_data(data)
        raw = "\n".join(parts)
        page_texts[pn] = norm(raw)
        all_parts.append(raw)

    full = norm("\n".join(all_parts))
    return page_texts, full


def build_page_texts_from_generic_pages(candidate: Dict[str, Any]) -> Optional[Tuple[Dict[int, str], str]]:
    pages = candidate.get("pages")
    if not isinstance(pages, list):
        return None

    page_texts: Dict[int, str] = {}
    all_parts: List[str] = []
    found_any = False

    for p in pages:
        if not isinstance(p, dict):
            continue

        pn = _to_int_page(p.get("page_number"))
        if pn is None:
            continue

        tmp: List[str] = []
        _gather_all_strings_generic(p, tmp)
        raw = "\n".join(tmp).strip()
        if raw:
            found_any = True
            page_texts[pn] = norm(raw)
            all_parts.append(raw)

    if not found_any:
        return None
    return page_texts, norm("\n".join(all_parts))


def build_candidate_index(candidate: Any) -> Tuple[Dict[int, str], str]:
    if isinstance(candidate, dict) and isinstance(candidate.get("raw_text"), str):
        return {}, norm(candidate.get("raw_text", ""))

    if looks_like_merged_glmocr(candidate):
        return build_page_texts_from_glm_merged(candidate)  # type: ignore[arg-type]

    if isinstance(candidate, dict):
        gen = build_page_texts_from_generic_pages(candidate)
        if gen is not None:
            return gen

    parts: List[str] = []
    _gather_all_strings_generic(candidate, parts)
    return {}, norm("\n".join(parts))


# ============================================================
# Evaluation (Overlap Recall)
# ============================================================

@dataclass
class MatchResult:
    path: str
    value: str
    page_number: Optional[int]
    matched: bool


def _sort_scope_keys(keys: List[str]) -> List[str]:
    nums = []
    has_doc = False
    others = []
    for k in keys:
        if k == "DOC":
            has_doc = True
        elif isinstance(k, str) and k.isdigit():
            nums.append(int(k))
        else:
            others.append(k)

    out = [str(x) for x in sorted(nums)]
    out.extend(sorted(others))
    if has_doc:
        out.append("DOC")
    return out


def evaluate_overlap_from_leaves(gt_leaves: List[GTLeaf], candidate: Any) -> Dict[str, Any]:
    page_texts_norm, full_text_norm = build_candidate_index(candidate)

    results: List[MatchResult] = []
    for leaf in gt_leaves:
        v_norm = norm(leaf.value)
        if not v_norm:
            continue

        if leaf.page_number is not None and leaf.page_number in page_texts_norm:
            hay = page_texts_norm[leaf.page_number]
        else:
            hay = full_text_norm

        matched = v_norm in hay
        results.append(MatchResult(leaf.path, leaf.value, leaf.page_number, matched))

    by_scope: Dict[str, Dict[str, int]] = {}
    for r in results:
        key = str(r.page_number) if r.page_number is not None else "DOC"
        if key not in by_scope:
            by_scope[key] = {"total": 0, "hit": 0}
        by_scope[key]["total"] += 1
        by_scope[key]["hit"] += 1 if r.matched else 0

    total = len(results)
    hit = sum(1 for r in results if r.matched)
    recall = (hit / total) if total else 0.0

    ordered_keys = _sort_scope_keys(list(by_scope.keys()))

    report = {
        "metric": "substring_overlap_recall",
        "total_fields_checked": total,
        "matched_fields": hit,
        "recall": round(recall, 6),
        "breakdown_by_page": {
            k: {
                "total": by_scope[k]["total"],
                "hit": by_scope[k]["hit"],
                "recall": round((by_scope[k]["hit"] / by_scope[k]["total"]) if by_scope[k]["total"] else 0.0, 6),
            }
            for k in ordered_keys
        },
    }
    return report


def evaluate_overlap(gt: Any, candidate: Any) -> Dict[str, Any]:
    gt_leaves = flatten_gt(gt, doc_type=_detect_doc_type_from_json(gt))
    return evaluate_overlap_from_leaves(gt_leaves, candidate)


# ============================================================
# Excel -> Leaves
# ============================================================

def _canon_section(doc_type: str, section: Any) -> str:
    s = str(section).strip().lower()
    if doc_type == "shipping_bill":
        if "shipping bill summary" in s:
            return "shippingbillsummary"
        if s == "header":
            return "header"
    if doc_type == "purchase_order":
        if "purchase" in s and "order" in s:
            return "purchaseorder"
    return _key_piece(section)


def shipping_bill_gt_excel_to_leaves(df: pd.DataFrame) -> List[GTLeaf]:
    cols = list(df.columns)
    if len(cols) < 4:
        return []

    c0, c1, c2, c3 = cols[0], cols[1], cols[2], cols[3]

    ff = df.copy()
    ff[c0] = ff[c0].ffill()
    ff[c1] = ff[c1].ffill()

    out: List[GTLeaf] = []
    for i, row in df.iterrows():
        section = ff.at[i, c0]
        group = ff.at[i, c1]
        r1, r2, r3 = row[c1], row[c2], row[c3]

        # Case A
        if (not _is_empty_cell(r2)) and (not _is_empty_cell(r3)):
            path = ".".join([p for p in [_canon_section("shipping_bill", section), _key_piece(group), _key_piece(r2)] if p])
            out.append(GTLeaf(path=path, value=leaf_to_string(r3), page_number=None))
            continue

        # Case B
        if (not _is_empty_cell(r1)) and (not _is_empty_cell(r2)):
            path = ".".join([p for p in [_canon_section("shipping_bill", section), _key_piece(r1)] if p])
            out.append(GTLeaf(path=path, value=leaf_to_string(r2), page_number=None))
            continue

    return out


def purchase_order_gt_excel_to_leaves(df: pd.DataFrame) -> List[GTLeaf]:
    cols = list(df.columns)
    if not cols:
        return []

    low = [str(c).lower() for c in cols]
    key_col = None
    val_col = None
    for i, c in enumerate(low):
        if c in {"key", "field", "path"}:
            key_col = cols[i]
        if c in {"value", "val"}:
            val_col = cols[i]

    if key_col is not None and val_col is not None:
        out: List[GTLeaf] = []
        for _, r in df.iterrows():
            k = r[key_col]
            v = r[val_col]
            if _is_empty_cell(k) or _is_empty_cell(v):
                continue
            out.append(GTLeaf(path=_key_piece(k).replace("_", "."), value=leaf_to_string(v), page_number=None))
        return out

    lcols = [c for c in cols if str(c).lower().startswith("l")]
    if "Value" in cols and lcols:
        out = []
        for _, r in df.iterrows():
            parts = []
            for c in lcols:
                if not _is_empty_cell(r[c]):
                    parts.append(_key_piece(r[c]))
            if not parts:
                continue
            v = r["Value"]
            if _is_empty_cell(v):
                continue
            out.append(GTLeaf(path=".".join(parts), value=leaf_to_string(v), page_number=None))
        return out

    if len(cols) >= 2:
        out = []
        for _, r in df.iterrows():
            k = r[cols[0]]
            v = r[cols[1]]
            if _is_empty_cell(k) or _is_empty_cell(v):
                continue
            out.append(GTLeaf(path=_key_piece(k).replace("_", "."), value=leaf_to_string(v), page_number=None))
        return out

    return []


def pred_excel_to_leaves(df: pd.DataFrame) -> List[GTLeaf]:
    cols = list(df.columns)
    if not cols:
        return []

    if "Path" in cols and "Value" in cols:
        out: List[GTLeaf] = []
        for _, r in df.iterrows():
            p = r["Path"]
            v = r["Value"]
            if _is_empty_cell(p):
                continue
            p_norm = _key_piece(p).replace("_", ".")
            out.append(GTLeaf(path=p_norm, value=leaf_to_string(v) if not _is_empty_cell(v) else "", page_number=None))
        return out

    lcols = [c for c in cols if str(c).lower().startswith("l")]
    if "Value" in cols and lcols:
        out = []
        for _, r in df.iterrows():
            parts = []
            for c in lcols:
                if not _is_empty_cell(r[c]):
                    parts.append(_key_piece(r[c]))
            if not parts:
                continue
            v = r["Value"]
            out.append(GTLeaf(path=".".join(parts), value=leaf_to_string(v) if not _is_empty_cell(v) else "", page_number=None))
        return out

    return []


def parse_gt(doc_type: str, fmt: str, payload: Any) -> List[GTLeaf]:
    if fmt == "json":
        return flatten_gt(payload, doc_type=doc_type)

    if not isinstance(payload, pd.DataFrame):
        return []

    if doc_type == "shipping_bill":
        return shipping_bill_gt_excel_to_leaves(payload)
    if doc_type == "purchase_order":
        return purchase_order_gt_excel_to_leaves(payload)
    return []


def parse_pred(fmt: str, payload: Any) -> List[GTLeaf]:
    if fmt == "json":
        detected = _detect_doc_type_from_json(payload, fallback="shipping_bill")
        return flatten_gt(payload, doc_type=detected)

    if not isinstance(payload, pd.DataFrame):
        return []

    return pred_excel_to_leaves(payload)


# ============================================================
# Alignment Excel
# ============================================================

def _sig(s: str) -> str:
    toks = re.split(r"[.\[\]_/:\-()\s=]+", s or "")
    normt = []
    for t in toks:
        t = t.strip()
        if not t:
            continue
        normt.append(_key_piece(t))
    normt = [x for x in normt if x]
    return "|".join(sorted(set(normt)))


def eligible_for_near_match(gt_raw: Any, pred_raw: Any) -> bool:
    """
    Generic, schema-agnostic gate for fuzzy matching.
    Only allow NEAR_MATCH for real text-like values.
    """
    if _is_empty_cell(gt_raw) or _is_empty_cell(pred_raw):
        return False

    gt_n = norm(gt_raw)
    pr_n = norm(pred_raw)

    if not gt_n or not pr_n:
        return False

    # tiny values should never be fuzzy-matched
    if len(gt_n) < 4 or len(pr_n) < 4:
        return False

    # exact numeric / identifier-like strings should be exact-only
    if is_numeric_like(gt_raw) and is_numeric_like(pred_raw):
        return False

    if looks_like_identifier(gt_raw) or looks_like_identifier(pred_raw):
        return False

    # only allow text-like phrases
    if not is_text_like_for_fuzzy(gt_raw):
        return False
    if not is_text_like_for_fuzzy(pred_raw):
        return False

    return True


def _status_for_values(path: str, gt_raw: Any, pred_raw: Any, jaccard_thresh: float = 0.85) -> str:
    if _is_empty_cell(pred_raw):
        return "MISSING_VALUE"

    gt_m = display_value(path, gt_raw)
    pr_m = display_value(path, pred_raw)

    # exact normalized match
    if gt_m == pr_m:
        return "MATCH"

    # loose match after removing commas and hyphens
    if norm_loose(gt_m) == norm_loose(pr_m):
        return "MATCH"

    # numeric exact-only check
    if is_numeric_like(gt_raw) and is_numeric_like(pred_raw):
        if norm_numeric(gt_raw) == norm_numeric(pred_raw):
            return "MATCH"
        return "MISMATCH"

    # Fuzzy matching only for real text-like values
    if not eligible_for_near_match(gt_raw, pred_raw):
        return "MISMATCH"

    sim = jaccard_tokens(pred_raw, gt_raw)
    if contains_norm(pred_raw, gt_raw) or sim >= jaccard_thresh:
        return "NEAR_MATCH"

    return "MISMATCH"


def build_alignment_xlsx_bytes(gt_leaves: List[GTLeaf], pred_leaves: List[GTLeaf]) -> bytes:
    """
    2-way: GT vs Pred
    Writes NORMALIZED display values into Excel.
    """
    pred_exact: Dict[str, List[int]] = {}
    for i, pl in enumerate(pred_leaves):
        k = (pl.path or "").strip()
        if k:
            pred_exact.setdefault(k, []).append(i)

    pred_sig: Dict[str, List[int]] = {}
    for i, pl in enumerate(pred_leaves):
        p = pl.path or ""
        segs = [x for x in p.split(".") if x]
        sigs = set()
        sigs.add(_sig(p))
        if segs:
            sigs.add(_sig(segs[-1]))
        if len(segs) >= 2:
            sigs.add(_sig(".".join(segs[-2:])))
        for s in sigs:
            if s:
                pred_sig.setdefault(s, []).append(i)

    used: set[int] = set()
    rows: List[Tuple[str, str, str, str, str]] = []

    for gl in gt_leaves:
        gk = (gl.path or "").strip()
        gv_raw = gl.value

        best_i: Optional[int] = None
        cands: List[int] = list(pred_exact.get(gk, []))

        if not cands:
            g_last = gk.split(".")[-1] if gk else ""
            gsig = _sig(g_last) or _sig(gk)
            cands = list(pred_sig.get(gsig, []))

        if cands:
            best_score = -1e9
            for ci in cands:
                score = 0.0
                if ci in used:
                    score -= 0.25

                pk = pred_leaves[ci].path or ""
                pv = pred_leaves[ci].value

                if pk.strip() == gk:
                    score += 4.0

                if not _is_empty_cell(pv) and norm_loose(display_value(gk, pv)) == norm_loose(display_value(gk, gv_raw)):
                    score += 2.0

                score -= abs(len(pk) - len(gk)) / 2000.0

                if score > best_score:
                    best_score = score
                    best_i = ci

        if best_i is None:
            rows.append((gk, "", display_value(gk, gv_raw), "", "MISSING_KEY"))
            continue

        used.add(best_i)
        pk = (pred_leaves[best_i].path or "").strip()
        pv_raw = pred_leaves[best_i].value

        st = _status_for_values(gk, gv_raw, pv_raw, jaccard_thresh=0.85)
        rows.append((gk, pk, display_value(gk, gv_raw), display_value(gk, pv_raw), st))

    wb = Workbook()
    ws = wb.active
    ws.title = "GT_vs_Pred"
    ws.append(["GT_Key(Hierarchy)", "Pred_Key(Hierarchy)", "GT_Value", "Pred_Value", "Status"])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="center", wrap_text=True)

    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    orange = PatternFill("solid", fgColor="FFD966")

    for r in rows:
        ws.append(list(r))

    for i in range(2, ws.max_row + 1):
        st = ws.cell(row=i, column=5).value
        if st == "MATCH":
            fill = green
        elif st == "NEAR_MATCH":
            fill = orange
        elif st == "MISMATCH":
            fill = yellow
        else:
            fill = red

        for j in (1, 3):
            ws.cell(row=i, column=j).alignment = center

        for j in (2, 4, 5):
            ws.cell(row=i, column=j).fill = fill
            ws.cell(row=i, column=j).alignment = center

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 16

    ws2 = wb.create_sheet("Legend")
    ws2.append(["Color", "Status", "Meaning"])
    for c in ws2[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    legend_rows = [
        ("Green", "MATCH", "GT value equals Pred value after normalization / loose punctuation handling"),
        ("Orange", "NEAR_MATCH", "Text-like containment OR Jaccard>=0.85"),
        ("Yellow", "MISMATCH", "Key matched but values differ"),
        ("Red", "MISSING_KEY / MISSING_VALUE", "No matched pred key OR pred value empty"),
    ]
    for row in legend_rows:
        ws2.append(list(row))

    for i in range(2, 6):
        status = ws2.cell(row=i, column=2).value
        if status == "MATCH":
            fill = green
        elif status == "NEAR_MATCH":
            fill = orange
        elif status == "MISMATCH":
            fill = yellow
        else:
            fill = red
        for j in range(1, 4):
            ws2.cell(row=i, column=j).fill = fill
            ws2.cell(row=i, column=j).alignment = center

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 58
    ws2.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_alignment_3way_xlsx_bytes(
    gt_leaves: List[GTLeaf],
    my_leaves: List[GTLeaf],
    gpt_leaves: List[GTLeaf],
    my_label: str = "MyModel",
    gpt_label: str = "GPT",
) -> bytes:
    """
    3-way: GT vs My vs GPT
    Writes NORMALIZED display values into Excel for GT/My/GPT values.
    Colors My columns by My status, GPT columns by GPT status.
    """

    def _index_pred(pred_leaves: List[GTLeaf]) -> Tuple[Dict[str, List[int]], Dict[str, List[int]]]:
        pred_exact: Dict[str, List[int]] = {}
        for i, pl in enumerate(pred_leaves):
            k = (pl.path or "").strip()
            if k:
                pred_exact.setdefault(k, []).append(i)

        pred_sig: Dict[str, List[int]] = {}
        for i, pl in enumerate(pred_leaves):
            p = pl.path or ""
            segs = [x for x in p.split(".") if x]
            sigs = set()
            sigs.add(_sig(p))
            if segs:
                sigs.add(_sig(segs[-1]))
            if len(segs) >= 2:
                sigs.add(_sig(".".join(segs[-2:])))
            for s in sigs:
                if s:
                    pred_sig.setdefault(s, []).append(i)

        return pred_exact, pred_sig

    def _best_match_for_gt(
        gl: GTLeaf,
        pred_leaves: List[GTLeaf],
        pred_exact: Dict[str, List[int]],
        pred_sig: Dict[str, List[int]],
        used: set[int],
    ) -> Tuple[str, Any, str]:
        """
        Returns (pred_key, pred_raw_value, status)
        """
        gk = (gl.path or "").strip()
        gv_raw = gl.value

        best_i: Optional[int] = None
        best_score = -1e9

        cands: List[int] = list(pred_exact.get(gk, []))

        if not cands:
            g_last = gk.split(".")[-1] if gk else ""
            gsig = _sig(g_last) or _sig(gk)
            cands = list(pred_sig.get(gsig, []))

        for ci in cands:
            score = 0.0

            pred_obj = pred_leaves[ci]
            pk = (pred_obj.path or "").strip()
            pv = pred_obj.value

            if ci in used:
                score -= 0.25

            if pk == gk:
                score += 4.0

            if not _is_empty_cell(pv):
                if norm_loose(display_value(gk, pv)) == norm_loose(display_value(gk, gv_raw)):
                    score += 2.0

            score -= abs(len(pk) - len(gk)) / 2000.0

            if score > best_score:
                best_score = score
                best_i = ci

        if best_i is None:
            return ("", "", "MISSING_KEY")

        used.add(best_i)

        pred_obj = pred_leaves[best_i]
        pk = (pred_obj.path or "").strip()
        pv_raw = pred_obj.value

        st = _status_for_values(gk, gv_raw, pv_raw, jaccard_thresh=0.85)
        return (pk, pv_raw, st)

    my_exact, my_sig = _index_pred(my_leaves)
    gpt_exact, gpt_sig = _index_pred(gpt_leaves)

    used_my: set[int] = set()
    used_gpt: set[int] = set()

    rows: List[Tuple[str, str, str, str, str, str, str, str]] = []

    for gl in gt_leaves:
        gk = (gl.path or "").strip()
        gv_raw = gl.value

        my_k, my_v_raw, my_st = _best_match_for_gt(gl, my_leaves, my_exact, my_sig, used_my)
        gpt_k, gpt_v_raw, gpt_st = _best_match_for_gt(gl, gpt_leaves, gpt_exact, gpt_sig, used_gpt)

        rows.append(
            (
                gk,
                my_k,
                gpt_k,
                display_value(gk, gv_raw),
                display_value(gk, my_v_raw),
                display_value(gk, gpt_v_raw),
                my_st,
                gpt_st,
            )
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "GT_vs_My_vs_GPT"

    ws.append(
        [
            "GT_Key(Hierarchy)",
            f"{my_label}_Pred_Key(Hierarchy)",
            f"{gpt_label}_Pred_Key(Hierarchy)",
            "GT_Value",
            f"{my_label}_Pred_Value",
            f"{gpt_label}_Pred_Value",
        ]
    )

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(vertical="center", wrap_text=True)

    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    orange = PatternFill("solid", fgColor="FFD966")

    def _fill_for_status(st: str) -> PatternFill:
        if st == "MATCH":
            return green
        if st == "NEAR_MATCH":
            return orange
        if st == "MISMATCH":
            return yellow
        return red

    for (gk, my_k, gpt_k, gt_disp, my_disp, gpt_disp, my_st, gpt_st) in rows:
        ws.append([gk, my_k, gpt_k, gt_disp, my_disp, gpt_disp])
        r = ws.max_row

        for j in (1, 4):
            ws.cell(row=r, column=j).alignment = center

        my_fill = _fill_for_status(my_st)
        for j in (2, 5):
            ws.cell(row=r, column=j).fill = my_fill
            ws.cell(row=r, column=j).alignment = center

        gpt_fill = _fill_for_status(gpt_st)
        for j in (3, 6):
            ws.cell(row=r, column=j).fill = gpt_fill
            ws.cell(row=r, column=j).alignment = center

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 28

    ws2 = wb.create_sheet("Legend")
    ws2.append(["Color", "Status", "Meaning"])
    for c in ws2[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    legend_rows = [
        ("Green", "MATCH", "GT value equals Pred value after normalization / loose punctuation handling"),
        ("Orange", "NEAR_MATCH", "Text-like containment OR Jaccard>=0.85"),
        ("Yellow", "MISMATCH", "Key matched but values differ"),
        ("Red", "MISSING_KEY / MISSING_VALUE", "No matched pred key OR pred value empty"),
    ]
    for row in legend_rows:
        ws2.append(list(row))

    for i in range(2, 6):
        status = ws2.cell(row=i, column=2).value
        fill = _fill_for_status(status)
        for j in range(1, 4):
            ws2.cell(row=i, column=j).fill = fill
            ws2.cell(row=i, column=j).alignment = center

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 58
    ws2.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()